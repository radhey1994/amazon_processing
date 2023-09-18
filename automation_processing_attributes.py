import openpyxl

wb = openpyxl.load_workbook("Automation_Excel_Reference.xlsx")
sheet = wb['Raw Sheet']

wb1 = openpyxl.load_workbook("output.xlsx")
sheet1 = wb1['Sheet1']
title_array = []
for row_no in range(2, sheet.max_row+1):
    product_details = sheet.cell(row=row_no, column=7).value
    products_list = product_details.split("|||")
    print(len(products_list))
    for product in products_list:
        if "=" in product:
            title, value = product.split("=")[0], product.split("=")[1]
            print(title, value)
            if title not in title_array:
                title_array.append(title)
        else:
            print("No Equal in ", product)
print(title_array)

for i in range(len(title_array)):
    sheet1.cell(row=1, column=i+1).value = title_array[i]

wb1.save("output1.xlsx")

for row_no in range(2, sheet.max_row+1):
    product_details = sheet.cell(row=row_no, column=7).value
    products_list = product_details.split("|||")
    for product in products_list:
        if "=" in product:
            title, value = product.split("=")[0], product.split("=")[1]
            for i in range(1, sheet1.max_column+1):
                if sheet1.cell(row=1, column=i).value == title:
                    sheet1.cell(row=row_no, column=i).value = value

wb1.save("attributes.xlsx")
wb1.close()