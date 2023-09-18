import openpyxl
import difflib

wb = openpyxl.load_workbook("Automation_Excel_Reference.xlsx")
sheet = wb['Raw Sheet']
sheet2 = wb['Category Sheet']
wb1 = openpyxl.load_workbook("output.xlsx")
sheet1 = wb1['Sheet1']

sheet1.cell(row=1, column=1).value = "Main Word"
sheet1.cell(row=1, column=2).value = "Other Words"
sheet1.cell(row=1, column=3).value = "Title"
sheet1.cell(row=1, column=4).value = "Category Mapping"
count = 2
for rn in range(2, sheet2.max_row+1):
    search_string = sheet2.cell(row=rn, column=1).value
    category = sheet2.cell(row=rn, column=2).value
    other_words = sheet2.cell(row=rn, column=3).value
    for row_no in range(2, sheet.max_row+1):
        cell_value = sheet.cell(row=row_no, column=2).value
        search_string1 = search_string.split()
        if all(x in cell_value for x in search_string1):
            print("Found 1\t", search_string, "******", cell_value, "***********", category)
            sheet1.cell(row=count, column=1).value = search_string
            sheet1.cell(row=count, column=3).value = cell_value
            sheet1.cell(row=count, column=4).value = category
            count += 1
        else:
            if other_words is not None:
                if any(y in cell_value for y in other_words.split(",")):
                    print("Found 2\t", other_words, "******", cell_value, "***********", category)
                    sheet1.cell(row=count, column=2).value = other_words
                    sheet1.cell(row=count, column=3).value = cell_value
                    sheet1.cell(row=count, column=4).value = category
                    count += 1

wb1.save("category.xlsx")
wb1.close()

