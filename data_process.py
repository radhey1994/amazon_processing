import openpyxl
import re
import json


input_wb = openpyxl.load_workbook("Automation_Inputs.xlsx")
input_wb_sheet = input_wb['Sheet1']
SHIPPING_TEMPLATE = input_wb_sheet.cell(row=2, column=4).value
ATTRIBUTES_LIST = []

for row_no in range(2, input_wb_sheet.max_row+1):
    attribute = input_wb_sheet.cell(row=row_no, column=5).value
    ATTRIBUTES_LIST.append(attribute)

f = open('package_weight_mappings.json')
PACKAGE_WEIGHT_MAPPING = json.load(f)
f.close()

SOURCE_FILE_NAME = input_wb_sheet.cell(row=2, column=1).value
source_wb = openpyxl.load_workbook(SOURCE_FILE_NAME)
source_wb_raw_sheet = source_wb['Sheet1']

CATEGORY_SHEET_FILE_NAME = input_wb_sheet.cell(row=2, column=3).value
category_wb = openpyxl.load_workbook(CATEGORY_SHEET_FILE_NAME)
category_wb_sheet = category_wb['Sheet1']

DESTINATION_FILE_NAME = input_wb_sheet.cell(row=2, column=2).value
destination_wb = openpyxl.load_workbook("Automation_Output_Template.xlsx")
destination_wb_sheet = destination_wb['Sheet1']

for i in range(len(ATTRIBUTES_LIST)):
    destination_wb_sheet.cell(row=1, column=19+i).value = ATTRIBUTES_LIST[i]

destination_wb.save(DESTINATION_FILE_NAME)
destination_wb.close()


def fetch_product_weight(product_title, product_details):
    """Method to Generate Item Weight from Product Title with below steps:
        1. Match Regex in title and try to find the weight.
        2. If not found in title, then get it from product details column
    """
    product_details_dict = {}
    products_details_list = product_details.split("|||")
    for product_detail in products_details_list:
        try:
            product_details_dict[product_detail.split("=")[0]] = product_detail.split("=")[1]
        except:
            product_details_dict[product_detail.split("=")[0]] = ""
    match = re.findall(r'([0-9,.]+( Ml| ML| ml| G| KG| kg| Kg|Grams|ml|G|ML|g|kg|KG|Kg))', product_title)
    if match:
        return match[0][0]
    else:
        return product_details_dict['Item Weight']


def fetch_images_urls(image_links):
    """Method to Generate comma separated list of image links with below steps:
        1. Split the images link with || as seperator
        2. join all the links with comma generated from split.
    """
    image_links = image_links.split("||")
    links_list = ""
    for link in image_links:
        links_list += link + ", "
    return links_list[:-1]


def assign_category(product_title):
    """Method to assign product category with below steps:
        1. Search for Main Word from Category Sheet in product title.
        2. If Main Word is present in title then map the category present in same row.
        3. If Main Word is not present in title then try matching each word in other words and assign the category.
    """
    for category_sheet_row_no in range(2, category_wb_sheet.max_row + 1):
        main_word = category_wb_sheet.cell(row=category_sheet_row_no, column=1).value
        category = category_wb_sheet.cell(row=category_sheet_row_no, column=2).value
        other_words = category_wb_sheet.cell(row=category_sheet_row_no, column=3).value
        main_word_split_list = main_word.split()
        if all(word in product_title for word in main_word_split_list):
            correct_category = category
            break
        else:
            if other_words is not None:
                if any(y in product_title for y in other_words.split(",")):
                    correct_category = category
                    break
            else:
                correct_category = ""
    return correct_category


def filter_brand_name(brand_name):
    """Method to filter brand name with below steps:
        1. check if specified list of words is present in brand name
        2. Remove which ever word exist and let rest be the same.
    """
    brand_name = brand_name.split(" ")
    not_wanted_list = ['Visit', 'the', 'Store']
    for word in not_wanted_list:
        if word in brand_name:
            brand_name.remove(word)
    filtered_brand_name = " ".join(brand_name)
    return filtered_brand_name


def calculate_package_weight(product_weight):
    """Method to calculate package weight below steps:
        1. From Product Weight, Get the Numeric Value and Unit Separately
        2. Now Based on Mapping, Add the value to Numeric Value
    """
    numeric = re.findall(r'[\d.]+', product_weight)
    string = re.findall(r'[a-zA-Z]+', product_weight)
    if len(numeric) > 1:
        weight = float(str(numeric[0]) + "." + str(numeric[1]))
    else:
        weight = float(numeric[0])
    add_weight = weight + PACKAGE_WEIGHT_MAPPING[string[0]]
    package_weight = str(add_weight) + " " + string[0]
    return package_weight


def fill_product_attributes(product_details, row_no):
    """Method to get product attributes from product details with below steps:
        1. Split the product details by |||.
        2. Now from split list, split each item with = and assign key and value to an empty dict.
    """
    attribute_and_values_mapping = {}
    attributes_value_list = product_details.split("|||")
    for attribute_and_value in attributes_value_list:
        if "=" in attribute_and_value:
            attribute, value = attribute_and_value.split("=")[0], attribute_and_value.split("=")[1]
            attribute_and_values_mapping[attribute] = value
    for k, v in attribute_and_values_mapping.items():
        if k in ATTRIBUTES_LIST:
            for i in range(1, destination_wb_sheet.max_column+1):
                if destination_wb_sheet.cell(row=1, column=i).value == k:
                    destination_wb_sheet.cell(row=row_no, column=i).value = v


for row_no in range(2, source_wb_raw_sheet.max_row+1):
    # Reading Data from Source File
    ASIN = source_wb_raw_sheet.cell(row=row_no, column=1).value
    PRODUCT_TITLE = source_wb_raw_sheet.cell(row=row_no, column=2).value
    BRAND = source_wb_raw_sheet.cell(row=row_no, column=3).value
    PRICE = source_wb_raw_sheet.cell(row=row_no, column=4).value
    DESCRIPTION = source_wb_raw_sheet.cell(row=row_no, column=6).value
    PRODUCT_DETAILS = source_wb_raw_sheet.cell(row=row_no, column=7).value
    PRODUCT_AVAILABILITY = source_wb_raw_sheet.cell(row=row_no, column=8).value
    PRODUCT_URL = source_wb_raw_sheet.cell(row=row_no, column=10).value
    IMAGES = source_wb_raw_sheet.cell(row=row_no, column=12).value
    BULLET_POINT1 = source_wb_raw_sheet.cell(row=row_no, column=13).value
    BULLET_POINT2 = source_wb_raw_sheet.cell(row=row_no, column=14).value
    BULLET_POINT3 = source_wb_raw_sheet.cell(row=row_no, column=15).value
    BULLET_POINT4 = source_wb_raw_sheet.cell(row=row_no, column=16).value
    BULLET_POINT5 = source_wb_raw_sheet.cell(row=row_no, column=17).value
    # Processing Data which is read from source file
    PRODUCT_WEIGHT = fetch_product_weight(PRODUCT_TITLE, PRODUCT_DETAILS)
    IMAGE_URLS = fetch_images_urls(IMAGES)
    PRODUCT_CATEGORY = assign_category(PRODUCT_TITLE)
    BRAND_NAME = filter_brand_name(BRAND)
    SKU = BRAND_NAME + "_" + str(row_no-1)
    VENDORID = BRAND_NAME
    PACKAGE_WEIGHT = calculate_package_weight(PRODUCT_WEIGHT)
    # Write Data to Destination File
    destination_wb_sheet.cell(row=row_no, column=1).value = ASIN
    destination_wb_sheet.cell(row=row_no, column=2).value = SKU
    destination_wb_sheet.cell(row=row_no, column=3).value = PRODUCT_TITLE
    destination_wb_sheet.cell(row=row_no, column=4).value = DESCRIPTION
    destination_wb_sheet.cell(row=row_no, column=5).value = PRODUCT_AVAILABILITY
    destination_wb_sheet.cell(row=row_no, column=6).value = PRODUCT_WEIGHT
    destination_wb_sheet.cell(row=row_no, column=7).value = PRICE
    destination_wb_sheet.cell(row=row_no, column=8).value = PRODUCT_CATEGORY
    destination_wb_sheet.cell(row=row_no, column=9).value = BRAND_NAME
    destination_wb_sheet.cell(row=row_no, column=10).value = PACKAGE_WEIGHT
    destination_wb_sheet.cell(row=row_no, column=11).value = BULLET_POINT1
    destination_wb_sheet.cell(row=row_no, column=12).value = BULLET_POINT2
    destination_wb_sheet.cell(row=row_no, column=13).value = BULLET_POINT3
    destination_wb_sheet.cell(row=row_no, column=14).value = BULLET_POINT4
    destination_wb_sheet.cell(row=row_no, column=15).value = BULLET_POINT5
    destination_wb_sheet.cell(row=row_no, column=16).value = IMAGES
    destination_wb_sheet.cell(row=row_no, column=17).value = SHIPPING_TEMPLATE
    destination_wb_sheet.cell(row=row_no, column=18).value = VENDORID
    fill_product_attributes(PRODUCT_DETAILS, row_no)

destination_wb.save(DESTINATION_FILE_NAME)
destination_wb.close()