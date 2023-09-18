import openpyxl
import difflib

wb = openpyxl.load_workbook("Automation_Excel_Reference.xlsx")
sheet = wb['Raw Sheet']
sheet2 = wb['Category Sheet']
wb1 = openpyxl.load_workbook("output.xlsx")
sheet1 = wb1['Sheet1']

for rn in range(2, sheet.max_row+1):
    image_links = sheet.cell(row=rn, column=12).value
    image_links = image_links.split("||")
    link1 = ""
    for link in image_links:
        link1 += link+", "
    sheet1.cell(row=rn-1, column=1).value = link1[:-1]


wb1.save("images.xlsx")
wb1.close()

