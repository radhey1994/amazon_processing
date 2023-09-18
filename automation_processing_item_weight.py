import openpyxl
import re

wb = openpyxl.load_workbook("Automation_Excel_Reference.xlsx")
sheet = wb['Raw Sheet']
wb1 = openpyxl.load_workbook("output.xlsx")
sheet1 = wb1['Sheet1']

title_array = []
sheet1.cell(row=1,column=1).value = "Product Title"
sheet1.cell(row=1, column=2).value = "Product Weight"
for row_no in range(2, sheet.max_row+1):
    product_details_dict = {}
    product_title = sheet.cell(row=row_no, column=2).value
    product_details = sheet.cell(row=row_no, column=7).value
    products_list = product_details.split("|||")
    print(products_list)
    for p in products_list:
        try:
            product_details_dict[p.split("=")[0]] = p.split("=")[1]
        except:
            product_details_dict[p.split("=")[0]] = ""
    match = re.findall(r'([0-9,.]+( Ml| ML| ml| G|Grams|ml|G|ML|g))', product_title)
    try:
        print(product_title,"\t", match[0][0])
        sheet1.cell(row=row_no,column=1).value = product_title
        sheet1.cell(row=row_no,column=2).value = match[0][0]
    except:
        print(product_title)
        sheet1.cell(row=row_no, column=1).value = product_title
        sheet1.cell(row=row_no, column=2).value = product_details_dict['Item Weight']

wb1.save("item_weight.xlsx")
wb1.close()

